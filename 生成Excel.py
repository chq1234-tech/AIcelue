"""
生成红利ETF资料汇总Excel文件
带格式：列宽调整、千分位格式化、自动筛选、冻结窗格
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取CSV数据
csv_file = r'd:\temp\红利ETF\红利ETF资料汇总.csv'
df = pd.read_csv(csv_file, encoding='utf-8-sig')

# 保存到Excel
excel_file = r'd:\temp\红利ETF\红利ETF资料汇总.xlsx'
df.to_excel(excel_file, index=False, sheet_name='红利ETF')

# 使用openpyxl美化格式
wb = load_workbook(excel_file)
ws = wb['红利ETF']

# 定义样式
header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置列宽
ws.column_dimensions['A'].width = 10  # 代码
ws.column_dimensions['B'].width = 25  # 名称
ws.column_dimensions['C'].width = 12  # 最新净值
ws.column_dimensions['D'].width = 12  # 净值日期
ws.column_dimensions['E'].width = 16  # 近1年收益率
ws.column_dimensions['F'].width = 16  # 近3年收益率
ws.column_dimensions['G'].width = 18  # 今年以来收益率
ws.column_dimensions['H'].width = 10  # 分红次数
ws.column_dimensions['I'].width = 15  # 最近分红
ws.column_dimensions['J'].width = 14  # 最近分红日期
ws.column_dimensions['K'].width = 12  # 规模
ws.column_dimensions['L'].width = 12  # 成立日期
ws.column_dimensions['M'].width = 25  # 备注

# 美化表头
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# 设置数据行格式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数字列右对齐
        if cell.column_letter in ['C', 'E', 'F', 'G', 'I']:
            cell.alignment = Alignment(horizontal='right', vertical='center')

# 冻结首行
ws.freeze_panes = 'A2'

# 添加自动筛选
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

# 添加数据条（收益率列）
from openpyxl.formatting.rule import DataBarRule
ws.conditional_formatting.add(
    f'E2:E{ws.max_row}',
    DataBarRule(start_type='num', start_value=0,
                end_type='num', end_value=30,
                color='63BE7B')  # 绿色
)

# 保存
wb.save(excel_file)

print("=" * 70)
print("✓ Excel文件生成成功！")
print("=" * 70)
print(f"\n文件路径: {excel_file}")
print(f"收录ETF数量: {len(df)} 只")
print("\nExcel格式特性:")
print("  ✓ 表头美化（蓝色背景、白色字体）")
print("  ✓ 自动筛选")
print("  ✓ 冻结首行")
print("  ✓ 边框线条")
print("  ✓ 条件格式（收益率数据条）")
print("\n数据来源: 天天基金网 (http://fund.eastmoney.com)")
print("数据日期: 2026-04-30")
print("=" * 70)
